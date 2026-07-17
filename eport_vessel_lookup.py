"""
ePort Saigon Newport - Tra cứu thông tin tàu/chuyến (v3)
Tự động: Khu vực = Cát Lái, nhập tên tàu, lấy thông tin tàu/chuyến

Tính năng:
  - Tìm nhiều tàu cùng lúc (song song) → nhanh
  - Tương tác qua DevExtreme API (ổn định, chính xác)
  - Hiển thị bảng tổng hợp + chi tiết
  - Xuất kết quả ra file Excel đầy đủ, định dạng đẹp

Yêu cầu:
    pip install selenium webdriver-manager colorama tabulate openpyxl
"""

import sys
import time
import json
from concurrent.futures import ThreadPoolExecutor, as_completed

# tabulate giúp in bảng đẹp; nếu chưa cài thì dùng bộ kẻ bảng tự viết
try:
    from tabulate import tabulate
    _HAS_TABULATE = True
except ImportError:
    _HAS_TABULATE = False

# openpyxl de xuat Excel; neu chua cai thi tat chuc nang xuat
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False

try:
    from selenium import webdriver
    from selenium.webdriver.common.by import By
    from selenium.webdriver.common.keys import Keys
    from selenium.webdriver.support.ui import WebDriverWait, Select
    from selenium.webdriver.support import expected_conditions as EC
    # Chrome
    from selenium.webdriver.chrome.options import Options as ChromeOptions
    from selenium.webdriver.chrome.service import Service as ChromeService
    # Edge
    from selenium.webdriver.edge.options import Options as EdgeOptions
    from selenium.webdriver.edge.service import Service as EdgeService
    from webdriver_manager.chrome import ChromeDriverManager
    from webdriver_manager.microsoft import EdgeChromiumDriverManager
    from colorama import init, Fore, Style
    init(autoreset=True)
except ImportError as _imp_err:
    print("\n[!] Thieu thu vien. Chay lenh sau de cai:")
    print("    pip install selenium webdriver-manager colorama tabulate openpyxl")
    print(f"    (chi tiet: {_imp_err})\n")
    sys.exit(1)


URL = "https://eport.saigonnewport.com.vn/Ships"
PAGE_TIMEOUT = 25
EL_TIMEOUT = 15

# Trinh duyet dang dung: "edge", "chrome", hoac "auto" (tu thu)
BROWSER = "auto"


def c_header(t): print(Fore.CYAN + Style.BRIGHT + t)
def c_ok(t):     print(Fore.GREEN + t)
def c_warn(t):   print(Fore.YELLOW + t)
def c_err(t):    print(Fore.RED + t)
def c_info(t):   print(Fore.WHITE + t)
def line():      print(Fore.BLUE + "-" * 64)


_CHROME_PATH = None
_EDGE_PATH = None


def _common_args(opts):
    """Cac tham so chung cho Chrome & Edge (deu la Chromium)."""
    opts.add_argument("--no-sandbox")
    opts.add_argument("--disable-dev-shm-usage")
    opts.add_argument("--disable-gpu")
    opts.add_argument("--disable-blink-features=AutomationControlled")
    opts.add_experimental_option("excludeSwitches", ["enable-automation"])
    opts.add_experimental_option("useAutomationExtension", False)
    opts.add_argument("--window-size=1400,900")
    opts.add_experimental_option(
        "prefs", {"profile.managed_default_content_settings.images": 2}
    )
    opts.add_argument("--blink-settings=imagesEnabled=false")
    opts.add_argument(
        "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
    )
    return opts


def _make_edge(headless):
    global _EDGE_PATH
    opts = EdgeOptions()
    if headless:
        opts.add_argument("--headless=new")
    _common_args(opts)
    if _EDGE_PATH is None:
        _EDGE_PATH = EdgeChromiumDriverManager().install()
    driver = webdriver.Edge(service=EdgeService(_EDGE_PATH), options=opts)
    return driver


def _make_chrome(headless):
    global _CHROME_PATH
    opts = ChromeOptions()
    if headless:
        opts.add_argument("--headless=new")
    _common_args(opts)
    if _CHROME_PATH is None:
        _CHROME_PATH = ChromeDriverManager().install()
    driver = webdriver.Chrome(service=ChromeService(_CHROME_PATH), options=opts)
    return driver


def create_driver(headless: bool = True):
    """Tao driver theo BROWSER. 'auto' = thu Edge truoc roi Chrome."""
    order = []
    if BROWSER == "edge":
        order = [("Edge", _make_edge)]
    elif BROWSER == "chrome":
        order = [("Chrome", _make_chrome)]
    else:  # auto
        order = [("Edge", _make_edge), ("Chrome", _make_chrome)]

    last_err = None
    for name, maker in order:
        try:
            driver = maker(headless)
            driver.execute_script(
                "Object.defineProperty(navigator,'webdriver',{get:()=>undefined})"
            )
            driver.set_page_load_timeout(PAGE_TIMEOUT)
            return driver
        except Exception as ex:
            last_err = ex
            continue
    raise RuntimeError(
        f"Khong mo duoc trinh duyet nao (Edge/Chrome). Loi cuoi: {last_err}"
    )


JS_INSPECT_FORM = r"""
function describe(el){
  return {
    tag: el.tagName.toLowerCase(),
    type: el.type || '',
    id: el.id || '',
    name: el.name || '',
    placeholder: el.placeholder || '',
    cls: el.className || '',
    value: el.value || '',
    visible: !!(el.offsetWidth || el.offsetHeight || el.getClientRects().length)
  };
}
var out = {inputs:[], selects:[], buttons:[], labels:[]};
document.querySelectorAll('input').forEach(e=>out.inputs.push(describe(e)));
document.querySelectorAll('select').forEach(e=>out.selects.push(describe(e)));
document.querySelectorAll('button, input[type=button], input[type=submit]').forEach(e=>{
  var d = describe(e); d.text = (e.innerText||e.value||'').trim(); out.buttons.push(d);
});
document.querySelectorAll('label').forEach(e=>out.labels.push((e.innerText||'').trim()));
return JSON.stringify(out);
"""


# ──────────────────────────────────────────────
# DevExtreme (dx) — trang ePort dung DevExtreme widgets
# Cach dung tin cay nhat: goi dx API qua JavaScript
# ID thuc te tren trang:
#   slbSiteId      = Khu vuc (SelectBox)   <-- KHOANH DO
#   txtVesselName  = Ten tau (TextBox)
#   devextreme2    = nut Tim kiem (Button)
#   grid-Vessels   = bang ket qua (DataGrid)
# ──────────────────────────────────────────────

# JS: chon Khu vuc = Cat Lai bang dx SelectBox API
# Cat Lai co SITE_ID="CTL", valueExpr="SITE_ID", displayExpr="SITE_NAME"
JS_SELECT_SITE = r"""
try {
  var el = document.getElementById('slbSiteId');
  if (!el) return 'NO_ELEMENT';
  if (typeof DevExpress === 'undefined') return 'NO_DEVEXPRESS';
  var inst = DevExpress.ui.dxSelectBox.getInstance(el)
          || (window.jQuery && jQuery('#slbSiteId').dxSelectBox('instance'));
  if (!inst) return 'NO_INSTANCE';

  // Cach chinh: set value = "CTL" (vi valueExpr la SITE_ID)
  // Dung 3-tham-so de DevExtreme phat su kien onValueChanged
  try {
    inst.option('value', 'CTL');
  } catch(e1) {}

  // Kiem tra da set chua
  var cur = inst.option('value');
  if (cur === 'CTL') {
    // Goi tay onValueChanged neu trang co handler rieng
    try {
      var h = inst.option('onValueChanged');
      if (h) h({value:'CTL', previousValue:null, component:inst, element:el});
    } catch(e2){}
    return 'OK:CTL';
  }

  // Fallback: duyet dataSource tim theo SITE_NAME = "Cat Lai"
  var ds = inst.option('dataSource');
  var items = [];
  try {
    if (ds && ds.store && ds.store()._array) items = ds.store()._array;
    else if (ds && ds.items && typeof ds.items === 'function') items = ds.items();
    else if (Array.isArray(ds)) items = ds;
  } catch(e3){}

  var target = null;
  for (var i=0;i<items.length;i++){
    var nm = (items[i].SITE_NAME || '').trim();
    if (nm === 'Cát Lái') { target = items[i]; break; }
  }
  if (target){
    inst.option('value', target.SITE_ID);
    return 'OK:' + target.SITE_ID + ' (' + target.SITE_NAME + ')';
  }

  return 'FAIL: value sau khi set = ' + JSON.stringify(cur) +
         ', items=' + items.length;
} catch(e){ return 'ERR:' + e.message; }
"""

# JS: nhap ten tau vao dx TextBox
JS_SET_VESSEL = r"""
try {
  var el = document.getElementById('txtVesselName');
  if (!el) return 'NO_ELEMENT';
  var inst = DevExpress.ui.dxTextBox.getInstance(el)
          || (window.jQuery && jQuery('#txtVesselName').dxTextBox('instance'));
  if (!inst) return 'NO_INSTANCE';
  inst.option('value', arguments[0]);
  return 'OK';
} catch(e){ return 'ERR:' + e.message; }
"""

# JS: bam nut Tim kiem (dx Button)
JS_CLICK_SEARCH = r"""
try {
  var el = document.getElementById('devextreme2');
  if (el){
    var inst = DevExpress.ui.dxButton.getInstance(el)
            || (window.jQuery && jQuery('#devextreme2').dxButton('instance'));
    if (inst){
      var h = inst.option('onClick');
      if (h){ h({}); return 'OK:dxButton.onClick'; }
    }
    // fallback: click DOM
    var content = el.querySelector('.dx-button-content') || el;
    content.click();
    return 'OK:domClick';
  }
  // fallback: tim theo text
  function noAccent(s){return (s||'').normalize('NFD').replace(/[\u0300-\u036f]/g,'').replace(/đ/g,'d').toLowerCase().trim();}
  var btns = document.querySelectorAll('.dx-button, button, input[type=button]');
  for (var b of btns){
    if (noAccent(b.innerText||b.value) === 'tim kiem'){ b.click(); return 'OK:textMatch'; }
  }
  return 'NOT_FOUND';
} catch(e){ return 'ERR:' + e.message; }
"""

# JS: doc du lieu tu dx DataGrid (grid-Vessels) qua API
JS_READ_GRID = r"""
try {
  var el = document.getElementById('grid-Vessels');
  if (!el) return JSON.stringify({error:'NO_GRID'});
  var inst = DevExpress.ui.dxDataGrid.getInstance(el)
          || (window.jQuery && jQuery('#grid-Vessels').dxDataGrid('instance'));
  if (!inst) return JSON.stringify({error:'NO_INSTANCE'});

  var rowsData = inst.getVisibleRows().map(function(r){ return r.data; });
  // Lay ten cot (caption) va dataField
  var cols = inst.getVisibleColumns()
                 .filter(function(c){ return c.dataField; })
                 .map(function(c){ return {field: c.dataField, caption: c.caption || c.dataField}; });
  return JSON.stringify({rows: rowsData, columns: cols});
} catch(e){ return JSON.stringify({error: e.message}); }
"""


def inspect_form(driver) -> dict:
    return json.loads(driver.execute_script(JS_INSPECT_FORM))


def select_cat_lai(driver):
    """Chon Khu vuc = Cat Lai qua dx SelectBox API (id=slbSiteId).
    Thu nhieu lan va xac minh gia tri da set thuc su."""
    verify_js = r"""
    try {
      var el = document.getElementById('slbSiteId');
      var inst = DevExpress.ui.dxSelectBox.getInstance(el)
              || (window.jQuery && jQuery('#slbSiteId').dxSelectBox('instance'));
      return inst ? inst.option('value') : null;
    } catch(e){ return 'ERR'; }
    """
    for attempt in range(3):
        try:
            driver.execute_script(JS_SELECT_SITE)
        except Exception:
            pass
        time.sleep(0.4)
        try:
            cur = driver.execute_script(verify_js)
            if cur == "CTL":
                return True
        except Exception:
            pass
        time.sleep(0.4)
    return False


def set_vessel_name(driver, vessel_name):
    """Nhap ten tau qua dx TextBox API (id=txtVesselName)."""
    try:
        result = driver.execute_script(JS_SET_VESSEL, vessel_name)
        return (result or "").startswith("OK")
    except Exception:
        return False


def click_search(driver):
    """Bam nut Tim kiem bang JS (dang tin cay hon XPath)."""
    # Cach 1: JS quet toan bo nut
    try:
        result = driver.execute_script(JS_CLICK_SEARCH)
        if result and result.startswith("OK"):
            return True
    except Exception:
        pass

    # Cach 2: Selenium XPath (du phong)
    xpaths = [
        u"//button[normalize-space()='T\u00ecm ki\u1ebfm']",
        u"//button[contains(.,'T\u00ecm ki\u1ebfm')]",
        u"//input[@type='submit' or @type='button'][contains(@value,'T\u00ecm')]",
        u"//*[self::button or self::a][contains(.,'T\u00ecm ki\u1ebfm')]",
        "//button[contains(@class,'search') or contains(@class,'primary')]",
    ]
    for xp in xpaths:
        try:
            b = driver.find_element(By.XPATH, xp)
            driver.execute_script("arguments[0].scrollIntoView({block:'center'});", b)
            driver.execute_script("arguments[0].click();", b)
            return True
        except Exception:
            continue

    # Cach 3: nhan Enter trong o nhap tau (nhieu form tu submit)
    try:
        vin = driver.switch_to.active_element
        vin.send_keys(Keys.ENTER)
        return True
    except Exception:
        pass

    return False


JS_READ_TABLE = r"""
// ===== Cach 1: doc tu <table> chuan =====
function readStandardTable(){
  var tables = document.querySelectorAll('table');
  var best = null, bestScore = -1;
  tables.forEach(t=>{
    var cols = t.querySelectorAll('th, thead td').length;
    var rows = t.querySelectorAll('tr').length;
    var score = cols * 10 + rows;
    if(score > bestScore){ bestScore = score; best = t; }
  });
  if(!best) return null;

  var rows = [];
  best.querySelectorAll('tr').forEach(tr=>{
    if(tr.closest('thead')) return;
    var tds = tr.querySelectorAll('td');
    if(tds.length === 0) return;
    if(tr.querySelector('input')) return;   // bo dong search
    var cells = [], hasText = false;
    tds.forEach(td=>{
      var txt = (td.innerText || '').replace(/\s+/g,' ').trim();
      cells.push(txt);
      if(txt.length>0) hasText = true;
    });
    if(hasText && cells.length>=3) rows.push(cells);
  });
  return rows.length ? rows : null;
}

// ===== Cach 2: doc tu grid dang <div> (role=row / class chua 'row') =====
function readDivGrid(){
  // Tim cac phan tu co role='row' hoac class chua 'gridrow'/'dx-row'/'k-grid'
  var rowEls = document.querySelectorAll(
    '[role=row], .dx-data-row, .k-master-row, .rgRow, .ant-table-row, tr[data-uid]'
  );
  var rows = [];
  rowEls.forEach(r=>{
    if(r.querySelector('input')) return;
    var cellEls = r.querySelectorAll('[role=gridcell], [role=cell], td, .dx-data-cell, .rgCell');
    if(cellEls.length < 3) return;
    var cells = [], hasText = false;
    cellEls.forEach(c=>{
      var txt = (c.innerText || '').replace(/\s+/g,' ').trim();
      cells.push(txt);
      if(txt.length>0) hasText = true;
    });
    if(hasText) rows.push(cells);
  });
  return rows.length ? rows : null;
}

var rows = readStandardTable() || readDivGrid();

// Neu van khong co -> tra ve HTML tho cua vung ket qua de chan doan
if(!rows){
  var area = document.body;
  // tim vung co chu "Thong tin tau" hoac bang gan nhat
  var hint = Array.from(document.querySelectorAll('*')).find(
    e => (e.innerText||'').indexOf('Thong tin tau') >= 0 ||
         (e.innerText||'').indexOf('Th\u00f4ng tin t\u00e0u') >= 0
  );
  var raw = (hint ? hint.outerHTML : document.body.innerHTML);
  return JSON.stringify({rows: [], rawHTML: raw.substring(0, 4000)});
}

return JSON.stringify({rows: rows, rawHTML: ''});
"""


def read_results(driver, want_raw=False):
    """Doc ket qua tu dx DataGrid API; neu that bai thi fallback DOM."""
    records = []
    raw_html = ""

    # ── Cach 1: dx DataGrid API (chinh xac nhat) ──
    try:
        data = json.loads(driver.execute_script(JS_READ_GRID))
        if "error" not in data:
            rows = data.get("rows", [])
            cols = data.get("columns", [])
            # Map caption -> ten cot tieng Viet ro rang
            for row in rows:
                if not isinstance(row, dict):
                    continue
                rec = {}
                for col in cols:
                    field = col.get("field")
                    caption = (col.get("caption") or field or "")
                    # caption co the chua \n (xuong dong) -> gop thanh 1 dong
                    caption = caption.replace("\\n", " ").replace("\n", " ")
                    caption = " ".join(caption.split())
                    if field in row:
                        val = row[field]
                        rec[caption] = "" if val is None else str(val)
                # neu khong co columns metadata, lay het cac field
                if not rec:
                    for k, v in row.items():
                        rec[k] = "" if v is None else str(v)
                if any(str(v).strip() for v in rec.values()):
                    records.append(rec)
            if records:
                return (records, raw_html) if want_raw else records
    except Exception:
        pass

    # ── Cach 2: fallback doc DOM (phong khi API loi) ──
    try:
        data = json.loads(driver.execute_script(JS_READ_TABLE))
        rows = data.get("rows", [])
        raw_html = data.get("rawHTML", "")

        col_names = [
            u"Terminal (C\u1ea3ng)",
            u"Agent (\u0110\u1ea1i l\u00fd H\u00e3ng t\u00e0u)",
            u"Vessel (T\u00ean t\u00e0u)",
            "Voyage In-Out",
            u"Berth time ATB (C\u1eadp th\u1ef1c t\u1ebf)",
            u"Departure time (R\u1eddi c\u1ea3ng)",
        ]
        header_markers = ("terminal", "agent", "vessel", "voyage", "berth", "departure",
                          u"c\u1ea3ng", u"\u0111\u1ea1i l\u00fd", u"t\u00ean t\u00e0u",
                          u"th\u1eddi gian", u"m\u00e3 chuy\u1ebfn")
        for row in rows:
            joined = " ".join(row).lower()
            if sum(1 for m in header_markers if m in joined) >= 4:
                continue
            rec = {}
            for i, val in enumerate(row):
                key = col_names[i] if i < len(col_names) else f"Cot {i+1}"
                rec[key] = val
            if any(v.strip() for v in rec.values()):
                records.append(rec)
    except Exception:
        pass

    return (records, raw_html) if want_raw else records


def _wait_devextreme_ready(driver, timeout=12):
    """Cho DevExtreme va cac widget (slbSiteId, txtVesselName) khoi tao xong."""
    deadline = time.time() + timeout
    js = r"""
    try {
      if (typeof DevExpress === 'undefined') return false;
      var a = document.getElementById('slbSiteId');
      var b = document.getElementById('txtVesselName');
      var c = document.getElementById('devextreme2');
      return !!(a && b && c);
    } catch(e){ return false; }
    """
    while time.time() < deadline:
        try:
            if driver.execute_script(js):
                return True
        except Exception:
            pass
        time.sleep(0.4)
    return False


# JS: dem so dong trong dx DataGrid
JS_GRID_ROWCOUNT = r"""
try {
  var el = document.getElementById('grid-Vessels');
  if (!el) return -1;
  var inst = DevExpress.ui.dxDataGrid.getInstance(el)
          || (window.jQuery && jQuery('#grid-Vessels').dxDataGrid('instance'));
  if (!inst) return -1;
  return inst.totalCount();
} catch(e){ return -1; }
"""


def wait_for_results_ready(driver, timeout=EL_TIMEOUT):
    """Cho den khi dx DataGrid co du lieu (totalCount > 0) hoac trang bao xong."""
    deadline = time.time() + timeout
    while time.time() < deadline:
        try:
            # Uu tien: dem so dong trong grid qua API
            cnt = driver.execute_script(JS_GRID_ROWCOUNT)
            if isinstance(cnt, (int, float)) and cnt > 0:
                return True
        except Exception:
            pass
        try:
            # Du phong: trang co chu "Tim thay X chuyen tau"
            body = driver.find_element(By.TAG_NAME, "body").text.lower()
            if "chuy" in body and ("t\u00ecm th\u1ea5y" in body or "tim thay" in body):
                # co the la "tim thay 0" -> van cho them chut
                return True
        except Exception:
            pass
        time.sleep(0.4)
    return False


def lookup_one(vessel_name: str, headless: bool = True, debug: bool = False):
    driver = create_driver(headless)
    wait = WebDriverWait(driver, EL_TIMEOUT)
    try:
        driver.get(URL)
        wait.until(EC.presence_of_element_located((By.TAG_NAME, "input")))
        # cho DevExtreme load xong (widget can JS khoi tao)
        _wait_devextreme_ready(driver, 12)
        time.sleep(1.0)
        if debug:
            return {"_debug": inspect_form(driver)}

        ok_area = select_cat_lai(driver)
        time.sleep(0.5)

        ok_vessel = set_vessel_name(driver, vessel_name)
        time.sleep(0.4)

        # Bam Tim kiem (dx Button API)
        clicked = click_search(driver)

        # Cho ket qua: dx DataGrid co dong du lieu
        ready = wait_for_results_ready(driver, EL_TIMEOUT)

        # Neu sau timeout van chua thay ket qua -> thu bam Tim kiem lan nua
        if not ready:
            click_search(driver)
            wait_for_results_ready(driver, 8)

        time.sleep(1.2)   # cho grid render xong

        # Doc bang, thu nhieu lan (cho lau hon vi bang co the load cham)
        deadline = time.time() + 12
        records, raw_html = [], ""
        while time.time() < deadline:
            records, raw_html = read_results(driver, want_raw=True)
            if records:
                break
            time.sleep(0.6)

        result = {"vessel": vessel_name, "records": records}
        # Neu khong co record, kem theo chan doan + LUU HTML ra file
        if not records:
            count_msg = _extract_count_msg(driver)
            html_file = ""
            try:
                # Luu toan bo HTML trang ra file de phan tich
                full_html = driver.page_source
                safe = "".join(ch if ch.isalnum() else "_" for ch in vessel_name)[:20]
                html_file = f"debug_{safe}.html"
                with open(html_file, "w", encoding="utf-8") as f:
                    f.write(full_html)
            except Exception:
                pass
            result["_diag"] = {
                "selected_cat_lai": ok_area,
                "set_vessel_name": ok_vessel,
                "clicked_search": clicked,
                "page_says": count_msg,
                "html_saved_to": html_file,
                "raw_html_snippet": raw_html[:1500],
            }
        return result
    except Exception as ex:
        return {"vessel": vessel_name, "error": str(ex).splitlines()[0]}
    finally:
        driver.quit()


def _extract_count_msg(driver):
    """Lay dong 'Tim thay X chuyen tau' tu trang."""
    try:
        body = driver.find_element(By.TAG_NAME, "body").text
        for ln in body.splitlines():
            low = ln.lower()
            if "chuy" in low and ("th\u1ea5y" in low or "thay" in low or
                                  "khong" in low or "kh\u00f4ng" in low):
                return ln.strip()
    except Exception:
        pass
    return "(khong doc duoc)"


def lookup_many(vessels, headless=True, max_workers=3):
    results = {}
    workers = min(max_workers, len(vessels))
    with ThreadPoolExecutor(max_workers=workers) as pool:
        futures = {pool.submit(lookup_one, v, headless): v for v in vessels}
        for fut in as_completed(futures):
            v = futures[fut]
            try:
                results[v] = fut.result()
            except Exception as ex:
                results[v] = {"vessel": v, "error": str(ex)}
    return results


def _show_diag(res: dict):
    """In thong tin chan doan khi khong lay duoc du lieu."""
    diag = res.get("_diag")
    if not diag:
        return
    c_warn("    --- CHAN DOAN (gui phan nay cho nguoi ho tro) ---")
    c_info(f"    - Da chon Cat Lai: {diag.get('selected_cat_lai')}")
    c_info(f"    - Da nhap ten tau: {diag.get('set_vessel_name')}")
    c_info(f"    - Da bam Tim kiem: {diag.get('clicked_search')}")
    c_info(f"    - Trang bao: {diag.get('page_says')}")
    html_file = diag.get("html_saved_to", "")
    if html_file:
        import os
        full_path = os.path.abspath(html_file)
        c_ok(f"    - DA LUU HTML RA FILE: {full_path}")
        c_warn(f"      >> Gui file nay cho nguoi ho tro de phan tich!")
    snippet = diag.get("raw_html_snippet", "")
    if snippet:
        c_info("    - HTML vung ket qua (rut gon):")
        print(Fore.LIGHTBLACK_EX + "      " + snippet[:800].replace("\n", "\n      "))
    c_warn("    -------------------------------------------------")


def show_one(res: dict):
    vessel = res.get("vessel", "?")
    line()
    if "error" in res:
        c_err(f"  x {vessel}: {res['error']}")
        return
    records = res.get("records", [])
    if not records:
        c_warn(f"  ! {vessel}: Khong tim thay chuyen tau nao")
        _show_diag(res)
        return
    c_ok(f"  v {vessel}: tim thay {len(records)} chuyen")
    for i, rec in enumerate(records, 1):
        c_header(f"\n  >> Chuyen #{i}")
        for k, val in rec.items():
            if not val:
                continue
            if "Berth" in k or "ATB" in k or u"c\u1eadp" in k.lower():
                print(f"    {Fore.YELLOW}{k}:")
                print(f"        {Fore.GREEN + Style.BRIGHT}>> {val}")
            else:
                print(f"    {Fore.WHITE}{k}: {Fore.CYAN}{val}")


# ──────────────────────────────────────────────
# Hiển thị dạng BẢNG tổng hợp tất cả các tàu
# ──────────────────────────────────────────────
# Các cột muốn hiển thị (rút gọn tên cho vừa màn hình)
TABLE_COLS = [
    ("Tau tim", None),                 # tên tàu người dùng nhập
    ("Terminal", "Terminal"),
    ("Agent", "Agent"),
    ("Vessel", "Vessel"),
    ("Voyage", "Voyage"),
    ("Berth ATB (Cap thuc te)", "Berth"),   # cột quan trọng
    ("Departure (Roi cang)", "Departure"),
]


def _match_col(rec: dict, keyword: str):
    """Tìm giá trị trong record có key chứa keyword (không phân biệt hoa thường)."""
    if keyword is None:
        return None
    kw = keyword.lower()
    for k, v in rec.items():
        if kw in k.lower():
            return v
    return ""


def _wrap(text: str, width: int) -> list[str]:
    """Cắt text thành nhiều dòng để vừa độ rộng cột."""
    text = (text or "").strip()
    if not text:
        return [""]
    words = text.split()
    lines, cur = [], ""
    for w in words:
        if len(cur) + len(w) + 1 <= width:
            cur = (cur + " " + w).strip()
        else:
            if cur:
                lines.append(cur)
            cur = w
    if cur:
        lines.append(cur)
    return lines or [""]


def show_table(results: dict, vessels: list[str]):
    """Gom kết quả tất cả tàu thành 1 bảng tổng hợp."""
    rows = []
    for v in vessels:
        res = results.get(v, {})
        if "error" in res:
            rows.append([v, "", "", "", "", f"LOI: {res['error'][:30]}", ""])
            continue
        records = res.get("records", [])
        if not records:
            rows.append([v, "", "", "", "", "Khong tim thay", ""])
            continue
        for rec in records:
            row = []
            for label, kw in TABLE_COLS:
                if kw is None:
                    row.append(v)
                else:
                    row.append(_match_col(rec, kw) or "")
            rows.append(row)

    headers = [label for label, _ in TABLE_COLS]

    line()
    c_header("  BANG TONG HOP KET QUA")
    line()

    if _HAS_TABULATE:
        # Tô màu cột Berth ATB (cột index 5)
        table = tabulate(rows, headers=headers, tablefmt="fancy_grid",
                         maxcolwidths=[10, 8, 8, 14, 16, 18, 18])
        # tô màu dòng tiêu đề + cột Berth bằng cách in cả bảng màu xanh nhạt
        print(Fore.CYAN + table)
    else:
        _print_table_fallback(headers, rows)

    # In riêng phần Berth ATB cho nổi bật
    print()
    c_header("  >> BERTH TIME (ATB) - Thoi gian tau cap thuc te:")
    for v in vessels:
        res = results.get(v, {})
        recs = res.get("records", [])
        if "error" in res:
            c_err(f"     {v:<16} : LOI")
        elif not recs:
            c_warn(f"     {v:<16} : Khong tim thay")
        else:
            for rec in recs:
                atb = _match_col(rec, "Berth") or "(trong)"
                print(f"     {Fore.WHITE}{v:<16} : {Fore.GREEN + Style.BRIGHT}{atb}")

    # Neu co tau khong tim thay -> in chan doan de gui ho tro
    for v in vessels:
        res = results.get(v, {})
        if not res.get("records") and "error" not in res and res.get("_diag"):
            print()
            c_warn(f"  [{v}] khong co du lieu:")
            _show_diag(res)


def _print_table_fallback(headers, rows):
    """Kẻ bảng thủ công khi chưa cài tabulate (hỗ trợ wrap nhiều dòng)."""
    widths = [10, 8, 8, 14, 16, 18, 18]

    def make_sep(left, mid, right, fill="─"):
        return left + mid.join(fill * (w + 2) for w in widths) + right

    def make_row(cells):
        # mỗi cell có thể nhiều dòng -> wrap
        wrapped = [_wrap(str(c), widths[i]) for i, c in enumerate(cells)]
        height = max(len(w) for w in wrapped)
        out_lines = []
        for line_i in range(height):
            parts = []
            for i, w in enumerate(wrapped):
                seg = w[line_i] if line_i < len(w) else ""
                parts.append(" " + seg.ljust(widths[i]) + " ")
            out_lines.append("│" + "│".join(parts) + "│")
        return "\n".join(out_lines)

    print(Fore.CYAN + make_sep("┌", "┬", "┐"))
    print(Fore.CYAN + Style.BRIGHT + make_row(headers))
    print(Fore.CYAN + make_sep("├", "┼", "┤"))
    for r in rows:
        print(Fore.WHITE + make_row(r))
        print(Fore.CYAN + make_sep("├", "┼", "┤"))
    print(Fore.CYAN + make_sep("└", "┴", "┘"))


# ──────────────────────────────────────────────
# Xuat ket qua ra file Excel (day du tat ca cot)
# ──────────────────────────────────────────────
def export_to_excel(results: dict, vessels: list[str], filename: str = None):
    """Xuat tat ca ket qua ra 1 file Excel dinh dang dep.
    Tra ve duong dan file da luu, hoac None neu loi."""
    if not _HAS_OPENPYXL:
        c_err("  ! Chua cai openpyxl. Chay: pip install openpyxl")
        return None

    import os
    from datetime import datetime

    if not filename:
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"eport_ket_qua_{stamp}.xlsx"

    # Gom tat ca records, kem cot "Tau tim kiem" o dau
    all_rows = []          # moi phan tu la dict
    all_columns = []       # thu tu cot xuat hien
    for v in vessels:
        res = results.get(v, {})
        recs = res.get("records", [])
        if "error" in res:
            all_rows.append({"Tàu tìm kiếm": v, "Trạng thái": f"LỖI: {res['error']}"})
            continue
        if not recs:
            all_rows.append({"Tàu tìm kiếm": v, "Trạng thái": "Không tìm thấy"})
            continue
        for rec in recs:
            row = {"Tàu tìm kiếm": v}
            row.update(rec)
            all_rows.append(row)
            # cap nhat danh sach cot theo thu tu gap
            for k in row.keys():
                if k not in all_columns:
                    all_columns.append(k)

    # Dam bao cot "Tau tim kiem" dung dau
    if "Tàu tìm kiếm" in all_columns:
        all_columns.remove("Tàu tìm kiếm")
    all_columns = ["Tàu tìm kiếm"] + all_columns

    # ── Tao workbook ──
    wb = Workbook()
    ws = wb.active
    ws.title = "Thông tin tàu"

    # Style
    FONT = "Arial"
    header_fill = PatternFill("solid", start_color="1F4E78")   # xanh dam
    header_font = Font(name=FONT, bold=True, color="FFFFFF", size=11)
    title_font = Font(name=FONT, bold=True, size=14, color="1F4E78")
    sub_font = Font(name=FONT, italic=True, size=9, color="808080")
    cell_font = Font(name=FONT, size=10)
    berth_font = Font(name=FONT, size=10, bold=True, color="006100")  # xanh la
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    alt_fill = PatternFill("solid", start_color="F2F7FB")      # xanh nhat xen ke

    # ── Tieu de bao cao ──
    ncol = len(all_columns)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncol)
    c = ws.cell(row=1, column=1, value="KẾT QUẢ TRA CỨU THÔNG TIN TÀU/CHUYẾN — ePort Saigon Newport")
    c.font = title_font
    c.alignment = center

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncol)
    info = (f"Khu vực: Cát Lái  |  Số tàu tra cứu: {len(vessels)}  |  "
            f"Xuất lúc: {datetime.now().strftime('%H:%M %d/%m/%Y')}")
    c = ws.cell(row=2, column=1, value=info)
    c.font = sub_font
    c.alignment = center

    # ── Header bang (dong 4) ──
    HEADER_ROW = 4
    for j, col_name in enumerate(all_columns, start=1):
        cell = ws.cell(row=HEADER_ROW, column=j, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    # ── Du lieu ──
    berth_col_idx = None
    for j, col_name in enumerate(all_columns, start=1):
        if "Berth" in col_name or "ATB" in col_name or "cập thực tế" in col_name.lower():
            berth_col_idx = j
            break

    for i, row in enumerate(all_rows):
        r = HEADER_ROW + 1 + i
        for j, col_name in enumerate(all_columns, start=1):
            val = row.get(col_name, "")
            cell = ws.cell(row=r, column=j, value=val)
            cell.border = border
            cell.alignment = left if j > 1 else center
            # to mau cot Berth
            if j == berth_col_idx and val:
                cell.font = berth_font
            else:
                cell.font = cell_font
            # xen ke mau dong
            if i % 2 == 1:
                cell.fill = alt_fill

    # ── Do rong cot tu dong (theo do dai noi dung) ──
    for j, col_name in enumerate(all_columns, start=1):
        max_len = len(str(col_name))
        for row in all_rows:
            v = str(row.get(col_name, ""))
            # lay dong dai nhat trong cac dong con (do wrap)
            longest = max((len(s) for s in v.split("\n")), default=0)
            max_len = max(max_len, longest)
        width = min(max(max_len + 2, 12), 45)   # gioi han 12-45
        ws.column_dimensions[get_column_letter(j)].width = width

    # Chieu cao dong header
    ws.row_dimensions[HEADER_ROW].height = 30

    # Freeze: giu header khi cuon
    ws.freeze_panes = ws.cell(row=HEADER_ROW + 1, column=1)

    # Bo loc (auto filter) cho bang
    last_row = HEADER_ROW + len(all_rows)
    ws.auto_filter.ref = f"A{HEADER_ROW}:{get_column_letter(ncol)}{last_row}"

    # ── Luu file ──
    try:
        wb.save(filename)
        return os.path.abspath(filename)
    except PermissionError:
        c_err(f"  ! Khong luu duoc '{filename}'. File co the dang mo trong Excel?")
        return None
    except Exception as ex:
        c_err(f"  ! Loi khi luu Excel: {ex}")
        return None



def show_debug(res: dict):
    d = res["_debug"]
    line()
    c_header("  CHAN DOAN - Cau truc form that cua trang")
    line()
    c_info(f"  Labels: {d['labels']}")
    c_header("\n  INPUTS:")
    for i in d["inputs"]:
        if i["visible"]:
            print(f"    id={i['id']!r} name={i['name']!r} "
                  f"placeholder={i['placeholder']!r} class={i['cls'][:40]!r}")
    c_header("\n  SELECTS:")
    for s in d["selects"]:
        print(f"    id={s['id']!r} name={s['name']!r} class={s['cls'][:40]!r}")
    c_header("\n  BUTTONS:")
    for b in d["buttons"]:
        if b["visible"]:
            print(f"    text={b.get('text','')!r} id={b['id']!r} class={b['cls'][:40]!r}")
    line()
    c_warn("  -> Gui phan nay cho minh de cap nhat selector neu tra cuu van loi.")


def main():
    global BROWSER
    line()
    c_header("  ePort Saigon Newport - Tra cuu Tau/Chuyen (v2)")
    line()
    c_info("  Khu vuc: Cat Lai  |  Tim song song nhieu tau  |  Toi uu toc do")
    line()

    # Chon trinh duyet
    br = input(
        Fore.WHITE + "\n  Trinh duyet:\n"
        "    [1] Tu dong (thu Edge truoc, roi Chrome) - khuyen dung\n"
        "    [2] Edge\n"
        "    [3] Chrome\n"
        "  Chon (mac dinh 1): "
    ).strip()
    BROWSER = {"2": "edge", "3": "chrome"}.get(br, "auto")
    c_info(f"  -> Dung trinh duyet: {BROWSER.upper()}")

    mode = input(
        Fore.WHITE + "\n  Che do:\n"
        "    [1] An browser (nhanh nhat) - khuyen dung\n"
        "    [2] Hien browser (de xem trinh duyet thao tac)\n"
        "    [3] Chan doan (in cau truc form - chi khi bi loi)\n"
        "  Chon (mac dinh 1): "
    ).strip()
    headless = mode != "2"
    debug = mode == "3"
    print()
    if debug:
        c_info("  -> Dang do cau truc trang...")
        res = lookup_one("TEST", headless=True, debug=True)
        if "_debug" in res:
            show_debug(res)
        else:
            c_err(f"  Loi: {res.get('error')}")
        return
    while True:
        raw = input(
            Fore.WHITE +
            "  Nhap ten tau (nhieu tau cach nhau boi dau phay ',', 'q' thoat):\n  > "
        ).strip()
        if raw.lower() in ("q", "quit", "exit", "thoat"):
            c_ok("\n  Tam biet!\n")
            break
        if not raw:
            c_warn("  ! Vui long nhap it nhat 1 ten tau.\n")
            continue
        vessels = [v.strip() for v in raw.split(",") if v.strip()]
        print()
        c_info(f"  Dang tim {len(vessels)} tau: {', '.join(vessels)}")
        line()
        start = time.time()
        if len(vessels) == 1:
            results = {vessels[0]: lookup_one(vessels[0], headless=headless)}
        else:
            results = lookup_many(vessels, headless=headless, max_workers=3)
        elapsed = time.time() - start
        # Hien thi dang BANG tong hop
        show_table(results, vessels)
        line()
        c_info(f"  Tong thoi gian: {elapsed:.1f}s cho {len(vessels)} tau")
        print()

        # Tuy chon xem chi tiet tung tau
        detail = input(
            Fore.WHITE + "  Xem chi tiet day du? (Enter = bo qua, 'd' = xem): "
        ).strip().lower()
        if detail == "d":
            for v in vessels:
                show_one(results.get(v, {"vessel": v, "error": "Khong co ket qua"}))
            print()

        # Tuy chon xuat ra Excel
        has_data = any(results.get(v, {}).get("records") for v in vessels)
        if has_data:
            exp = input(
                Fore.WHITE + "  Xuat ket qua ra file Excel? (Enter = co, 'n' = khong): "
            ).strip().lower()
            if exp != "n":
                path = export_to_excel(results, vessels)
                if path:
                    c_ok(f"  ✓ Da xuat Excel: {path}")
            print()

        again = input(Fore.WHITE + "  Tim tiep? (Enter = co, q = thoat): ").strip()
        if again.lower() in ("q", "quit", "n", "no"):
            c_ok("\n  Tam biet!\n")
            break
        print()


if __name__ == "__main__":
    main()
